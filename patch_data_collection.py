import re

with open("development/data_collection.py", "r") as f:
    content = f.read()

search = """
    print("Attempting to load game list from Google Sheet...")
    google_sheet_url = "https://docs.google.com/spreadsheets/d/1B2S4kj4PY_U7W5daQyLbv1XvFIFm64o0lFv0Q4fxZIA/export?format=xlsx"

    today = datetime.date.today().isoformat()
    # Resolve the data directory dynamically relative to this script
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
    filepath = os.path.join(data_dir, f"{today}-gsheets.xlsx")

    # ensure data directory exists
    os.makedirs(data_dir, exist_ok=True)

    if not os.path.exists(filepath):
        print(f"File {filepath} not found. Downloading...")

        # Clean up old gsheets files in the data directory
        for f_name in os.listdir(data_dir):
            if f_name.startswith("20") and ("-gsheets.xlsx" in f_name or "-gsheets.csv" in f_name) and not f_name.startswith(today):
                try:
                    os.remove(os.path.join(data_dir, f_name))
                    print(f"Removed old cache file: {f_name}")
                except Exception as e:
                    print(f"Warning: Failed to remove old cache file {f_name}: {e}")

        try:
            response = requests.get(google_sheet_url)
            response.raise_for_status()
            with open(filepath, 'wb') as f:
                f.write(response.content)
            print(f"Downloaded successfully to {filepath}")
        except Exception as e:
            print(f"Failed to download Google Sheet: {e}")
            return pd.DataFrame()
    else:
        print(f"File {filepath} already exists. Using cached version.")
"""

replace = """
    import glob

    print("Attempting to load game list from Google Sheet...")
    google_sheet_url = "https://docs.google.com/spreadsheets/d/1B2S4kj4PY_U7W5daQyLbv1XvFIFm64o0lFv0Q4fxZIA/export?format=xlsx"

    today_date = datetime.date.today()
    today = today_date.isoformat()
    # Resolve the data directory dynamically relative to this script
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
    filepath = os.path.join(data_dir, f"{today}-gsheets.xlsx")

    # ensure data directory exists
    os.makedirs(data_dir, exist_ok=True)

    # Find the most recent file
    existing_files = glob.glob(os.path.join(data_dir, "20*-gsheets.xlsx")) + glob.glob(os.path.join(data_dir, "20*-gsheets.csv"))
    existing_files.sort(reverse=True)

    needs_scrape = True
    latest_file = None

    if existing_files:
        latest_file = existing_files[0]
        latest_filename = os.path.basename(latest_file)
        date_str_match = re.match(r"^(\d{4}-\d{2}-\d{2})-gsheets", latest_filename)
        if date_str_match:
            try:
                latest_date = datetime.date.fromisoformat(date_str_match.group(1))
                if (today_date - latest_date).days < 1:
                    needs_scrape = False
            except ValueError:
                pass

    if needs_scrape or os.environ.get("FORCE_SCRAPE") == "1":
        print(f"Initiating new scrape. Downloading to {filepath}...")
        try:
            response = requests.get(google_sheet_url)
            response.raise_for_status()

            # Save to temporary file first for validation
            temp_filepath = filepath + ".tmp"
            with open(temp_filepath, 'wb') as f:
                f.write(response.content)

            # Basic Validation
            import openpyxl
            wb = openpyxl.load_workbook(temp_filepath, data_only=True)
            sheet = wb.active

            # Find the header row (skipping 15 rows, so row 16 is header)
            # data starts at 17
            row_count = 0
            empty_names = 0
            for row_idx in range(17, sheet.max_row + 1):
                # Only check non-empty rows to count correctly
                # We consider row has content if column A (FROM) or B (TO) has something
                if sheet.cell(row=row_idx, column=1).value is not None:
                    row_count += 1
                    # check column F (NAME)
                    if sheet.cell(row=row_idx, column=6).value is None:
                        empty_names += 1

            if row_count < 100:
                print(f"Validation failed: Scrape contains too few rows ({row_count}).")
                os.remove(temp_filepath)
                # use latest if we failed
                filepath = latest_file if latest_file else filepath
            elif row_count > 0 and (empty_names / row_count) > 0.5:
                print(f"Validation failed: Too many empty names ({empty_names}/{row_count}).")
                os.remove(temp_filepath)
                filepath = latest_file if latest_file else filepath
            else:
                os.rename(temp_filepath, filepath)
                print(f"Downloaded and validated successfully to {filepath}")
                latest_file = filepath

                # Maintain rolling pair: keep today and the previous valid one, delete older ones
                all_files_now = glob.glob(os.path.join(data_dir, "20*-gsheets.xlsx")) + glob.glob(os.path.join(data_dir, "20*-gsheets.csv"))
                all_files_now.sort(reverse=True)
                # Keep top 2
                for old_f in all_files_now[2:]:
                    try:
                        os.remove(old_f)
                        print(f"Removed old cache file: {os.path.basename(old_f)}")
                    except Exception as e:
                        print(f"Warning: Failed to remove old cache file {old_f}: {e}")

        except Exception as e:
            print(f"Failed to download or validate Google Sheet: {e}")
            if os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                except Exception:
                    pass
            # If download fails, try to fallback to latest
            if latest_file:
                 print(f"Falling back to {latest_file}")
                 filepath = latest_file
            else:
                 return pd.DataFrame()
    else:
        print(f"Using cached version: {latest_file}.")
        filepath = latest_file

    if not filepath or not os.path.exists(filepath):
        print("No valid file path to read from.")
        return pd.DataFrame()

"""

if search in content:
    content = content.replace(search, replace)
    with open("development/data_collection.py", "w") as f:
        f.write(content)
    print("Patched successfully")
else:
    print("Search string not found")

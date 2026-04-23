import pandas as pd
import numpy as np
import re
import time
import json
from epicstore_api import EpicGamesStoreAPI, EGSException
# import requests                # web fallback
# from bs4 import BeautifulSoup  # used for web fallback

# verify API wrapper is working
try:
    api = EpicGamesStoreAPI()
    print("EpicGamesStoreAPI initialized successfully.")
except Exception as e:
    print(f"Failed to initialize EpicGamesStoreAPI: {e}")
    api = None

import os
import datetime
import requests

def scrape_gamelist():
    """
    Scrapes the list of free games and dates from the Google Sheet data source.

    Returns a pandas DataFrame with nice column names.
    """

    import glob

    print("Attempting to load game list from Google Sheet...")
    google_sheet_url = "https://docs.google.com/spreadsheets/d/1B2S4kj4PY_U7W5daQyLbv1XvFIFm64o0lFv0Q4fxZIA/export?format=xlsx"

    today_date = datetime.date.today()
    today = today_date.isoformat()
    # Resolve the data directory dynamically relative to this script
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
    filepath = os.path.join(data_dir, f"{today}-gsheets.xlsx")

    # ensure data directory exists
    os.makedirs(data_dir, exist_ok=True)

    # Find the most recent file
    existing_files = glob.glob(os.path.join(data_dir, "20*-gsheets.xlsx")) + glob.glob(os.path.join(data_dir, "20*-gsheets.csv"))
    existing_files.sort(reverse=True)

    needs_scrape = True
    latest_file = None

    if existing_files:
        latest_file = existing_files[0]
        latest_filename = os.path.basename(latest_file)
        date_str_match = re.match(r"^(\d{4}-\d{2}-\d{2})-gsheets", latest_filename)
        if date_str_match:
            try:
                latest_date = datetime.date.fromisoformat(date_str_match.group(1))
                if (today_date - latest_date).days < 1:
                    needs_scrape = False
            except ValueError:
                pass

    if needs_scrape or os.environ.get("FORCE_SCRAPE") == "1":
        print(f"Initiating new scrape. Downloading to {filepath}...")
        try:
            response = requests.get(google_sheet_url)
            response.raise_for_status()

            # Save to temporary file first for validation
            temp_filepath = filepath + ".tmp"
            with open(temp_filepath, 'wb') as f:
                f.write(response.content)

            # Basic Validation
            import openpyxl
            wb = openpyxl.load_workbook(temp_filepath, data_only=True)
            sheet = wb.active

            # Find the header row (skipping 15 rows, so row 16 is header)
            # data starts at 17
            row_count = 0
            empty_names = 0
            for row_idx in range(17, sheet.max_row + 1):
                # Only check non-empty rows to count correctly
                # We consider row has content if column A (FROM) or B (TO) has something
                if sheet.cell(row=row_idx, column=1).value is not None:
                    row_count += 1
                    # check column F (NAME)
                    if sheet.cell(row=row_idx, column=6).value is None:
                        empty_names += 1

            if row_count < 100:
                print(f"Validation failed: Scrape contains too few rows ({row_count}).")
                os.remove(temp_filepath)
                # use latest if we failed
                filepath = latest_file if latest_file else filepath
            elif row_count > 0 and (empty_names / row_count) > 0.5:
                print(f"Validation failed: Too many empty names ({empty_names}/{row_count}).")
                os.remove(temp_filepath)
                filepath = latest_file if latest_file else filepath
            else:
                os.rename(temp_filepath, filepath)
                print(f"Downloaded and validated successfully to {filepath}")
                latest_file = filepath

                # Maintain rolling pair: keep today and the previous valid one, delete older ones
                all_files_now = glob.glob(os.path.join(data_dir, "20*-gsheets.xlsx")) + glob.glob(os.path.join(data_dir, "20*-gsheets.csv"))
                all_files_now.sort(reverse=True)
                # Keep top 2
                for old_f in all_files_now[2:]:
                    try:
                        os.remove(old_f)
                        print(f"Removed old cache file: {os.path.basename(old_f)}")
                    except Exception as e:
                        print(f"Warning: Failed to remove old cache file {old_f}: {e}")

        except Exception as e:
            print(f"Failed to download or validate Google Sheet: {e}")
            if os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                except Exception:
                    pass
            # If download fails, try to fallback to latest
            if latest_file:
                 print(f"Falling back to {latest_file}")
                 filepath = latest_file
            else:
                 return pd.DataFrame()
    else:
        print(f"Using cached version: {latest_file}.")
        filepath = latest_file

    if not filepath or not os.path.exists(filepath):
        print("No valid file path to read from.")
        return pd.DataFrame()


    try:
        import openpyxl

        # We need to skip 15 rows of intro stuff
        gdf = pd.read_excel(filepath, engine='openpyxl', skiprows=15)

        # Extract colors via openpyxl before doing any pandas filtering that changes row indices
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        color_map = {
            'FFD0E0E3': 'regular',
            'FFB6D7A8': 'mystery',
            'FFFFF2CC': 'unusual',
            'FFF4CCCC': 'mobile'
        }

        colors = []
        # pd.read_excel with skiprows=15 reads row 16 as header, data starts at row 17.
        # So we iterate openpyxl from row 17 to len(gdf) + 16
        for row_idx in range(17, 17 + len(gdf)):
            cell = sheet.cell(row=row_idx, column=6) # 6th column is 'NAME'
            rgb = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None

            # Map rgb to known categories, default to 'unknown' if not matched
            category = color_map.get(rgb, 'unknown') if rgb and type(rgb) == str else 'unknown'
            colors.append(category)

        # Take the first 7 columns and rename them appropriately, making sure the 6th is 'Games'
        # so it matches expected schema. (Original sheet columns: FROM, TO, DAY, DAYS, TYPE, NAME, NOTES)
        gdf = gdf.iloc[:, 0:7].copy()
        gdf.columns = ['FROM', 'TO', 'DAY', 'DAYS', 'TYPE', 'Games', 'NOTES']

        # Assign the parsed colors as a new column
        gdf['COLOR_CATEGORY'] = colors

        # Slice from the first "next" and exclude trailing rows
        first_next_idx = gdf.index[gdf['TYPE'] == 'next'].tolist()
        if first_next_idx:
            start_idx = first_next_idx[0]
            gdf = gdf.loc[start_idx:].copy()

        # Drop the last row if it repeats the header
        if len(gdf) > 0 and gdf.iloc[-1]['FROM'] == 'FROM':
            gdf = gdf.iloc[:-1].copy()

        # Fall back to 'NOTES' if 'Games' is empty
        # Handled carefully: note the column was renamed from 'NAME / NOTES' to 'NOTES' and 'NAME' to 'Games' above!
        gdf['Games'] = gdf['Games'].fillna(gdf['NOTES'])

        # Forward fill missing dates and metadata
        gdf['FROM'] = gdf['FROM'].ffill()
        gdf['TO'] = gdf['TO'].ffill()
        gdf['DAY'] = gdf['DAY'].ffill()
        gdf['DAYS'] = gdf['DAYS'].ffill()

        # Filter out rows that are not valid games (future placeholders, dividers, or empty names)
        # Placeholder TYPE often is '*', valid ones might be empty or 'mobile', 'other'.
        # We also need to strip whitespace and drop nans in Games.
        gdf = gdf[~gdf['Games'].isna()].copy()
        gdf['Games'] = gdf['Games'].astype("string").str.strip()
        gdf = gdf[~gdf['Games'].isin(['', '-', 'nan'])]
        gdf = gdf[gdf['TYPE'] != '*']

        # We reset index for cleaner look
        gdf = gdf.reset_index(drop=True)

        print("Successfully loaded data from Google Sheet.")
        print(f"Columns found: {gdf.columns.tolist()}")
        print("First 5 rows of the loaded data:")
        print(gdf.head())

        return gdf
        
    except Exception as e:
        print(f"Failed to load data from {filepath}: {e}")
        return pd.DataFrame()

def get_game_details(game_title: str, delay: float = 1.5):
    """
    Fetches details from Epic Games Store API by title
    
    Returns closest match title's JSON, None if nothing found

    #TODO: return in a more structured format AND indicate if the match was exact or inferred
    """
    
    time.sleep(delay)
    print(f"Fetching details for: {game_title}...")
    try:
        # NOTE: The API wrapper's get_product might require a namespace/offer_id, which we don't have yet.
        search_results = api.fetch_store_games(keywords=str(game_title).strip(), count = 5)
        
        # if we get a response AND it matches the JSON format we expect, look at possible matches
        if search_results and search_results.get('data', {}).get('Catalog', {}).get('searchStore', {}).get('elements'):
            elements = search_results['data']['Catalog']['searchStore']['elements']
            
            best_match = None
            for element in elements:
                if str(element.get('title', '')).strip().lower() == str(game_title).strip().lower():
                    best_match = element
                    print(f"  Exact match found: {element.get('title')}")
                    break

            #TODO: add more extensive matching logic and waittime, following "archive/code/reviewget.py"
            # for now, just use the first result if no exact match
            
            if not best_match and elements:
                best_match = elements[0]
                print(f"  No exact match. Using first result: {best_match.get('title')}")
            
            if best_match:
                return best_match
            else:
                print(f"  No search results found for '{game_title}'.")
                return None
        else:
            print(f"  No search results structure for '{game_title}'.")
            return None
    except EGSException as e:
        print(f"  API Error fetching details for '{game_title}': {e}")
        return None
    except Exception as e:
        print(f"  Unexpected error fetching details for '{game_title}': {e}")
        return None

def main():
    """
    Main execution pipeline for Epic Games Store data collection.

    Retrieves the primary game list from Google Sheets, iterates through the targets,
    queries the Epic Games Store API for detailed product info (tags, price, etc.),
    and saves the enriched result to a CSV file.
    """
    games_df = scrape_gamelist()
    
    if games_df.empty:
        print("Could not load game data from sheet. Exiting.")
        return

    all_game_details = []
    # TESTING: only look at first 5 games
    games_to_process = games_df.head(5) 
    # games_to_process = games_df

    for index, row in games_to_process.iterrows():
        game_title = row['Games']
        if pd.isna(game_title) or str(game_title).strip() == "":
            print(f"Skipping row {index} due to empty game title.")
            continue
            
        details = get_game_details(game_title)
        
        if details:
            # Combine original row data with fetched details
            # This is a basic merge; more sophisticated merging might be needed
            combined_info = row.to_dict()
            combined_info['api_details'] = details # Store the whole API response for now
            # We'll refine what to extract from 'details' later (tags, price, etc.)
            # Example: combined_info['api_title'] = details.get('title')
            #          combined_info['api_description'] = details.get('description')
            #          combined_info['api_tags'] = details.get('tags') # Structure of tags needs checking
            all_game_details.append(combined_info)
        else:
            # Still add original info even if API fetch fails, marking it
            combined_info = row.to_dict()
            combined_info['api_details'] = None
            all_game_details.append(combined_info)
            
        print("--- Waiting before next API call ---")
        time.sleep(2) # RATE LIMITING: Wait for 2 seconds between API calls

    # Convert the list of dictionaries to a DataFrame
    processed_games_df = pd.DataFrame(all_game_details)
    
    print("\nProcessed Games DataFrame:")
    print(processed_games_df.head())
    
    # Save the processed data
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
    output_path = os.path.join(data_dir, "epic_games_data_with_api_details.csv")
    try:
        processed_games_df.to_csv(output_path, index=False)
        print(f"Successfully saved processed data to {output_path}")
    except Exception as e:
        print(f"Failed to save data: {e}")

# ensure imports do not auto-run main()
if __name__ == "__main__":
    main() 